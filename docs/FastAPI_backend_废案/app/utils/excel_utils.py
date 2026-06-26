"""Excel 处理工具"""
import io
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
from fastapi import HTTPException, status


def generate_score_template(assessment_items: List[Dict[str, Any]], class_students: List[Dict[str, str]]) -> bytes:
    """
    生成复合表头 Excel 成绩录入模板
    Row 1: 学号 | 姓名 | 考核点1 | 考核点2 | ...
    Row 2:      |       | 满分:100 / 目标:1-1 | 满分:80 / 目标:2-1 | ...
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩录入"

    # Styles
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    sub_font = Font(name="微软雅黑", size=10, bold=True, color="333333")
    sub_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Row 1: 主表头
    ws.cell(row=1, column=1, value="学号")
    ws.cell(row=1, column=2, value="姓名")
    for idx, item in enumerate(assessment_items, start=3):
        ws.cell(row=1, column=idx, value=item["name"])

    for col in range(1, len(assessment_items) + 3):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Row 2: 子表头
    for idx, item in enumerate(assessment_items, start=3):
        obj_code = item.get("objective_code", "")
        max_score = item.get("max_score", 0)
        ws.cell(row=2, column=idx, value=f"满分:{max_score} / 目标:{obj_code}")

    for col in range(1, len(assessment_items) + 3):
        cell = ws.cell(row=2, column=col)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Row 3+: 学生数据
    for row_idx, student in enumerate(class_students, start=3):
        ws.cell(row=row_idx, column=1, value=student.get("student_id", ""))
        ws.cell(row=row_idx, column=2, value=student.get("name", ""))
        for col in range(3, len(assessment_items) + 3):
            ws.cell(row=row_idx, column=col, value="")
            ws.cell(row=row_idx, column=col).border = thin_border

    # Freeze panes
    ws.freeze_panes = "C3"

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    import string
    for i in range(len(assessment_items)):
        col_letter = string.ascii_uppercase[i + 2] if i < 24 else "A" + string.ascii_uppercase[i - 24]
        ws.column_dimensions[col_letter].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def parse_score_excel(file_content: bytes, expected_items: List[Dict[str, Any]]) -> tuple:
    """
    解析上传的成绩 Excel，返回 (student_scores, errors)
    student_scores: List[{"student_id": str, "item_id": str, "score": float}]
    """
    item_name_map = {item["name"]: item["id"] for item in expected_items}
    item_max_scores = {item["id"]: float(item["max_score"]) for item in expected_items}

    try:
        df = pd.read_excel(io.BytesIO(file_content), header=None)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Excel 解析失败: {str(e)}")

    if df.empty or len(df) < 3:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel 文件格式不正确，至少需要 3 行数据")

    # Validate headers (Row 0)
    row0 = [str(v).strip() for v in df.iloc[0].tolist()]
    if row0[0] != "学号" or row0[1] != "姓名":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Excel 表头格式错误，第 1 列应为"学号"，第 2 列应为"姓名"')

    # Validate assessment item headers match exactly
    uploaded_items = row0[2:]
    missing_items = set(item_name_map.keys()) - set(uploaded_items)
    extra_items = set(uploaded_items) - set(item_name_map.keys())
    if missing_items or extra_items:
        error_msg = "考核点表头不匹配"
        if missing_items:
            error_msg += f"，缺少: {', '.join(missing_items)}"
        if extra_items:
            error_msg += f"，多余: {', '.join(extra_items)}"
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error_msg)

    # Parse score data
    scores = []
    errors = []
    for idx in range(2, len(df)):
        row = df.iloc[idx]
        student_id = str(row.iloc[0]).strip()
        student_name = str(row.iloc[1]).strip()
        if not student_id:
            continue
        for col_idx, item_name in enumerate(uploaded_items, start=2):
            item_id = item_name_map[item_name]
            raw_score = row.iloc[col_idx]
            try:
                score = float(raw_score) if pd.notna(raw_score) else None
            except (ValueError, TypeError):
                errors.append(f"第{idx + 1}行 考核点 {item_name} 成绩格式错误")
                continue
            if score is not None:
                max_score = item_max_scores[item_id]
                if score < 0 or score > max_score:
                    errors.append(f"第{idx + 1}行 考核点 {item_name} 成绩 {score} 超出范围 [0, {max_score}]")
                else:
                    scores.append({"student_id": student_id, "item_id": item_id, "score": score})

    return scores, errors


def generate_course_report_excel(objective_results: list, course_results: list, student_scores: list) -> bytes:
    """
    生成课程级《课程目标达成情况评价表》Excel（三个 Sheet）
    """
    wb = Workbook()

    # Sheet 1: 课程目标达成度
    ws1 = wb.active
    ws1.title = "课程目标达成度"
    ws1.cell(row=1, column=1, value="课程目标")
    ws1.cell(row=1, column=2, value="达成度 C_bar")
    for idx, result in enumerate(objective_results, start=2):
        ws1.cell(row=idx, column=1, value=result.get("objective_code", ""))
        ws1.cell(row=idx, column=2, value=result.get("C_bar", 0))

    # Sheet 2: 指标点达成度
    ws2 = wb.create_sheet("指标点达成度")
    ws2.cell(row=1, column=1, value="指标点")
    ws2.cell(row=1, column=2, value="达成度 E_k")
    for idx, result in enumerate(course_results, start=2):
        ws2.cell(row=idx, column=1, value=result.get("indicator_code", ""))
        ws2.cell(row=idx, column=2, value=result.get("E_k", 0))

    # Sheet 3: 学生明细
    ws3 = wb.create_sheet("学生明细")
    headers = ["学号", "姓名"]
    obj_codes = [r.get("objective_code", "") for r in objective_results]
    headers.extend([f"目标{c} 达成度" for c in obj_codes])
    for col_idx, header in enumerate(headers, start=1):
        ws3.cell(row=1, column=col_idx, value=header)
    for row_idx, student in enumerate(student_scores, start=2):
        ws3.cell(row=row_idx, column=1, value=student.get("student_id", ""))
        ws3.cell(row=row_idx, column=2, value=student.get("name", ""))
        for col_idx, obj_code in enumerate(obj_codes, start=3):
            ws3.cell(row=row_idx, column=col_idx, value=student.get(f"C_{obj_code}", 0))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def generate_program_report_excel(program_results: dict, indicator_detail: dict) -> bytes:
    """
    生成专业级穿透式台账 Excel
    """
    wb = Workbook()

    # Sheet 1: 专业达成度汇总
    ws1 = wb.active
    ws1.title = "专业达成度汇总"
    ws1.cell(row=1, column=1, value="指标点编号")
    ws1.cell(row=1, column=2, value="指标点描述")
    ws1.cell(row=1, column=3, value="达成度 G_k")
    for idx, (ind_id, data) in enumerate(program_results.items(), start=2):
        ws1.cell(row=idx, column=1, value=data.get("indicator_code", ""))
        ws1.cell(row=idx, column=2, value=data.get("indicator_description", ""))
        ws1.cell(row=idx, column=3, value=data.get("achievement_value", 0))

    # Sheet 2+: 每个指标点的支撑课程明细
    for ind_id, detail in indicator_detail.items():
        ws = wb.create_sheet(f"指标点-{ind_id}")
        ws.cell(row=1, column=1, value="课程")
        ws.cell(row=1, column=2, value="课程达成度 E_k")
        ws.cell(row=1, column=3, value="宏观权重 W")
        ws.cell(row=1, column=4, value="贡献值 Ek*W")
        for idx, row in enumerate(detail, start=2):
            ws.cell(row=idx, column=1, value=row.get("course_name", ""))
            ws.cell(row=idx, column=2, value=row.get("E_k", 0))
            ws.cell(row=idx, column=3, value=row.get("W", 0))
            ws.cell(row=idx, column=4, value=row.get("contribution", 0))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
